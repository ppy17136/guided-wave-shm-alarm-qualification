#!/usr/bin/env python3
"""Audit and index the 2024 healthy variable-temperature CFRP SEG-2 dataset."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import struct
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd


NUMBER_RE = re.compile(r"^1_0_(?P<number>\d+)_")


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(4 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def read_param_strings(handle) -> dict[str, str]:
    params: dict[str, str] = {}
    while True:
        raw = handle.read(2)
        if len(raw) != 2:
            raise EOFError("Unexpected EOF in SEG-2 parameter block")
        length = struct.unpack("<h", raw)[0]
        if length <= 0:
            break
        text = handle.read(length - 2).decode("latin-1", errors="replace").replace("\x00", " ")
        parts = text.strip().split(maxsplit=1)
        key = parts[0].lower().replace(".", "_") if parts else ""
        value = parts[1] if len(parts) > 1 else ""
        if key:
            params[key] = value
    return params


def read_sg2(path: Path, load_data: bool = True) -> dict:
    with path.open("rb") as handle:
        descriptor = handle.read(32)
        if len(descriptor) != 32:
            raise ValueError(f"Short SEG-2 header: {path}")
        values = struct.unpack("<HHHHBBBBBB18B", descriptor)
        file_descriptor_id = values[0]
        if file_descriptor_id != 0x3A55:
            raise ValueError(f"Invalid SEG-2 descriptor {file_descriptor_id:#x}: {path}")
        revision = values[1]
        pointer_block_size = values[2]
        trace_count = values[3]
        pointers = np.fromfile(handle, dtype="<u4", count=trace_count)
        handle.seek(32 + pointer_block_size)
        file_params = read_param_strings(handle)

        trace_headers = []
        traces = []
        for pointer in pointers:
            handle.seek(int(pointer))
            trace_descriptor = handle.read(32)
            if len(trace_descriptor) != 32:
                raise ValueError(f"Short trace descriptor: {path}")
            trace_values = struct.unpack("<HHIIB19B", trace_descriptor)
            trace_id = trace_values[0]
            if trace_id != 0x4422:
                raise ValueError(f"Invalid trace descriptor {trace_id:#x}: {path}")
            trace_header = {
                "size_of_block": trace_values[1],
                "following_data_bytes": trace_values[2],
                "samples": trace_values[3],
                "format_code": trace_values[4],
            }
            trace_text = read_param_strings(handle)
            position = handle.tell()
            handle.seek(position + ((4 - position % 4) % 4))
            if load_data:
                dtype = {
                    1: np.dtype("<i2"),
                    2: np.dtype("<i4"),
                    4: np.dtype("<f4"),
                    5: np.dtype("<f8"),
                }.get(trace_header["format_code"])
                if dtype is None:
                    raise ValueError(
                        f"Unsupported SEG-2 format {trace_header['format_code']}: {path}"
                    )
                signal = np.fromfile(handle, dtype=dtype, count=trace_header["samples"]).astype(
                    np.float64, copy=False
                )
                if signal.size != trace_header["samples"]:
                    raise ValueError(f"Short trace data: {path}")
                traces.append(signal)
            trace_headers.append({**trace_header, "text": trace_text})

    data = None
    if load_data:
        sample_counts = {trace.size for trace in traces}
        if len(sample_counts) != 1:
            raise ValueError(f"Inconsistent trace lengths in {path}")
        data = np.stack(traces, axis=0)
    return {
        "revision": revision,
        "pointer_block_size": pointer_block_size,
        "trace_count": trace_count,
        "file_params": file_params,
        "trace_headers": trace_headers,
        "data": data,
    }


def load_climate_mapping(path: Path, files: int) -> list[dict]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    header = [str(value) if value is not None else "" for value in next(iterator)]
    records = [dict(zip(header, row)) for row in iterator if row[0] is not None]
    # MATLAB script: table row k=11, then k=k+5. Python records are zero-based.
    indices = [10 + 5 * index for index in range(files)]
    mapped = []
    for index in indices:
        if index >= len(records):
            mapped.append(
                {
                    "timestamp": None,
                    "temperature_C": np.nan,
                    "temperature_program_C": np.nan,
                    "humidity_pct": np.nan,
                    "humidity_program_pct": np.nan,
                    "climate_record_zero_based": None,
                    "worksheet_row": None,
                    "climate_metadata_missing": True,
                }
            )
            continue
        record = records[index]
        mapped.append(
            {
                "timestamp": record["Date/Time"].isoformat(sep=" "),
                "temperature_C": float(record["TEMP_MEAS"]),
                "temperature_program_C": float(record["TEMP_PROG"]),
                "humidity_pct": float(record["RH_MEAS"]),
                "humidity_program_pct": float(record["RH_PROG"]),
                "climate_record_zero_based": index,
                "worksheet_row": index + 2,
                "climate_metadata_missing": False,
            }
        )
    return mapped


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--raw-root", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    sg2_files = sorted(
        args.raw_root.rglob("*.sg2"),
        key=lambda path: int(NUMBER_RE.match(path.name).group("number")),
    )
    if len(sg2_files) != 220:
        raise SystemExit(f"Expected 220 SG2 files, found {len(sg2_files)}")
    excel_files = [
        path
        for path in args.raw_root.rglob("USER*.xlsx")
        if not path.name.startswith("~$")
    ]
    if len(excel_files) != 1:
        raise SystemExit(f"Expected one climate workbook, found {len(excel_files)}")
    climate = load_climate_mapping(excel_files[0], len(sg2_files))

    rows = []
    shape_set = set()
    format_set = set()
    for sequence, (path, climate_record) in enumerate(zip(sg2_files, climate), start=1):
        parsed = read_sg2(path, load_data=True)
        data = parsed["data"]
        shape_set.add(tuple(data.shape))
        format_codes = sorted(
            {header["format_code"] for header in parsed["trace_headers"]}
        )
        format_set.update(format_codes)
        rows.append(
            {
                "sequence": sequence,
                "file_number": int(NUMBER_RE.match(path.name).group("number")),
                "filename": path.name,
                "bytes": path.stat().st_size,
                "sha256": file_sha256(path),
                "traces": int(data.shape[0]),
                "samples": int(data.shape[1]),
                "format_codes": ",".join(map(str, format_codes)),
                "finite": bool(np.isfinite(data).all()),
                "minimum": float(np.min(data)),
                "maximum": float(np.max(data)),
                "rms": float(np.sqrt(np.mean(data**2))),
                **climate_record,
            }
        )

    frame = pd.DataFrame(rows)
    frame.to_csv(args.output_dir / "composite_june2024_acquisition_index.csv", index=False)
    temperatures = frame.temperature_C.dropna().to_numpy()
    timestamps = pd.to_datetime(frame.timestamp, errors="coerce")
    intervals_seconds = timestamps.diff().dt.total_seconds().dropna().to_numpy()
    summary = {
        "schema_version": "p9e-composite-june2024-audit-v1",
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "source": "https://zenodo.org/records/19209079",
        "doi": "10.5281/zenodo.19209079",
        "license": "CC BY 4.0",
        "sg2_files": len(sg2_files),
        "sg2_total_bytes": int(frame.bytes.sum()),
        "shape_set": [list(shape) for shape in sorted(shape_set)],
        "format_codes": sorted(format_set),
        "all_finite": bool(frame.finite.all()),
        "file_numbers": [int(frame.file_number.min()), int(frame.file_number.max())],
        "temperature_C": {
            "minimum": float(temperatures.min()),
            "maximum": float(temperatures.max()),
            "median": float(np.median(temperatures)),
        },
        "humidity_pct": {
            "minimum": float(frame.humidity_pct.min()),
            "maximum": float(frame.humidity_pct.max()),
            "median": float(frame.humidity_pct.median()),
        },
        "interval_seconds": {
            "minimum": float(intervals_seconds.min()),
            "maximum": float(intervals_seconds.max()),
            "median": float(np.median(intervals_seconds)),
        },
        "checks": {
            "exactly_220_files": len(sg2_files) == 220,
            "climate_metadata_for_219_files": int((~frame.climate_metadata_missing).sum()) == 219,
            "one_consistent_shape": len(shape_set) == 1,
            "forty_four_traces": shape_set and next(iter(shape_set))[0] == 44,
            "all_finite": bool(frame.finite.all()),
            "five_minute_median_mapping": bool(240 <= np.median(intervals_seconds) <= 360),
            "temperature_spans_at_least_45C": bool(
                temperatures.max() - temperatures.min() >= 45
            ),
        },
    }
    summary["passed"] = all(summary["checks"].values())
    (args.output_dir / "composite_june2024_audit.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8"
    )
    report = [
        "# CompositeTest June 2024 独立健康变温数据审计",
        "",
        f"- SG2文件：{summary['sg2_files']}",
        f"- 具有官方脚本气候映射的文件：{int((~frame.climate_metadata_missing).sum())}/220",
        f"- 一致形状：{summary['shape_set']}",
        f"- 温度范围：{summary['temperature_C']['minimum']:.3f}–{summary['temperature_C']['maximum']:.3f} ℃",
        f"- 湿度中位数：{summary['humidity_pct']['median']:.3f}%",
        f"- 采集间隔中位数：{summary['interval_seconds']['median']:.1f} s",
        f"- 全部有限数值：{'是' if summary['all_finite'] else '否'}",
        f"- 总体通过：{'是' if summary['passed'] else '否'}",
        "",
        "官方MATLAB索引只为前219个SG2提供气候记录；最后一个文件保留但标记为气候元数据缺失。",
        "该数据全程无损伤，只能用于独立健康温度迁移、误报和未知状态审计，不能用于损伤召回或AUROC确认。",
    ]
    (args.output_dir / "composite_june2024_audit_report.md").write_text(
        "\n".join(report) + "\n", encoding="utf-8"
    )
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
